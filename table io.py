"""
Writing result tables.

Every stage that produces results - detection, refinement and trains - writes the same pair of
files: a CSV that is always written, and an Excel copy on request. The code for that lives here
rather than in any one stage, so that no stage has to import another merely to write its output. An
earlier version kept it in detect_pulses.py and had the other two import it from there, which made
refinement and train-finding depend on the detector for a reason that has nothing to do with
detection, and broke outright whenever the copy of detect_pulses.py in use predated the helper.

openpyxl is imported only when an Excel file is actually requested, so the rest of the pipeline
runs without it installed.
"""

import csv


def save_table_csv(header, rows, path):

    """
    Writes a table as CSV.

    ----------

    Parameters:
        header (list) - column names.
        rows (list) - one list of values per row, in the same order as the header.
        path (str) - output file path.

    Returns:
        path (str) - the file written.
    """

    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    return path


def save_table_xlsx(header, rows, path, sheet="data"):

    """
    Writes a table as an Excel file, keeping numbers as numbers so the result can be sorted and
    filtered in a spreadsheet without further conversion. The header row is frozen and bold, and
    the columns are widened to fit their contents.

    ----------

    Parameters:
        header (list) - column names.
        rows (list) - one list of values per row, in the same order as the header.
        path (str) - output file path.
        sheet (str) - default 'data'. Worksheet name.

    Returns:
        path (str) - the file written, or None when openpyxl is not installed.
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("  openpyxl is not installed, so no Excel file was written "
              "(the CSV holds the same table)")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    def cell(v):

        """
        Numbers as numbers. The tables are assembled for a CSV, so a value that came from one is a
        string even when it is numeric; written straight to a sheet it would sort alphabetically
        and refuse to plot. Anything that is not a number is passed through unchanged, so labels,
        timestamps and empty fields survive intact.
        """

        if v is None or isinstance(v, (int, float)):
            return v
        if isinstance(v, str):
            t = v.strip()
            if t == "":
                return None
            try:
                return int(t) if t.lstrip("+-").isdigit() else float(t)
            except ValueError:
                return v
        return str(v)

    for row in rows:
        ws.append([cell(v) for v in row])

    for i, name in enumerate(header, start=1):
        longest = len(str(name))
        for row in rows:
            if i - 1 < len(row):
                longest = max(longest, len(str(row[i - 1])))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(longest + 2, 40)

    wb.save(path)
    return path
